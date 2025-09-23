import streamlit as st
import gspread
import pandas as pd
import io
import os
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
import datetime
import pytz
import time 

# ========== Google Sheet 設定 ==========
SHEET_NAME = "Project_Form"
WORKSHEET_NAME = "Python"
LOCK_SHEET_NAME = "Lock"
TAIWAN_TZ = pytz.timezone("Asia/Taipei")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

creds = Credentials.from_service_account_info(
    st.secrets["gcp_service_account"],
    scopes=SCOPES
)
client = gspread.authorize(creds)
sheet = client.open(SHEET_NAME).worksheet(WORKSHEET_NAME)

# 固定 Google Sheet 欄位順序
SHEET_HEADERS = [
    "Project_Number", "Sales_User", "ODM_Customers", "Brand_Customers", "Application_Purpose",
    "Project_Name", "Proposal_Date", "Product_Application", "Cooling_Solution",
    "Delivery_Location", "Sample_Date", "Sample_Qty", "Demand_Qty",
    "SI", "PV", "MV", "MP", "Spec_Type", "Update_Time",
    "Sales_Manager", "RD_Manager", "Apply_Date", "Applicant_Name",
    "Fill_Date", "Spec_Writer", "Sales_Review", "RD_Review", "Approval"
]

# ========== 使用者帳號密碼 ==========
USER_CREDENTIALS = {
    "sam@kipotec.com.tw": {"password": "Kipo-0926969586$$$", "name": "Sam"},
    "sale1@kipotec.com.tw": {"password": "Kipo-0917369466$$$", "name": "Vivian"},
    "sale5@kipotec.com.tw": {"password": "Kipo-0925698417$$$", "name": "Wendy"},
    "sale2@kipotec.com.tw": {"password": "Kipo-0905038111$$$", "name": "Lillian"},
}

# 優先順序
USER_PRIORITY = {"Sam": 1, "Vivian": 2, "Lillian": 3, "Wendy": 4}

# ========== Lock 機制 ==========
def open_lock_ws():
    sh = client.open(SHEET_NAME)
    try:
        ws_lock = sh.worksheet(LOCK_SHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        ws_lock = sh.add_worksheet(title=LOCK_SHEET_NAME, rows=10, cols=2)
        ws_lock.update("A1:B1", [["User", "Locked_Time"]])
    return ws_lock

def load_lock_df():
    ws_lock = open_lock_ws()
    data = ws_lock.get_all_records()
    df_lock = pd.DataFrame(data)
    if "User" not in df_lock.columns:
        df_lock["User"] = ""
    if "Locked_Time" not in df_lock.columns:
        df_lock["Locked_Time"] = ""
    return df_lock, ws_lock

def acquire_lock(username: str) -> (bool, str):
    df_lock, ws_lock = load_lock_df()
    active = df_lock[df_lock["User"] != ""]
    now = datetime.datetime.now(TAIWAN_TZ)   # ✅ 改成台灣時間

    if active.empty:
        ws_lock.append_row([username, now.strftime("%Y-%m-%d %H:%M:%S")])
        return True, ""

    current_user = active.iloc[0]["User"]
    lock_time = datetime.datetime.strptime(
    active.iloc[0]["Locked_Time"], "%Y-%m-%d %H:%M:%S"
    ).replace(tzinfo=TAIWAN_TZ)

    time_diff = (now - lock_time).total_seconds()

    if current_user == username:
        return True, ""

    if time_diff <= 3:
        current_pri = USER_PRIORITY.get(current_user, 99)
        new_pri = USER_PRIORITY.get(username, 99)
        if new_pri < current_pri:
            ws_lock.update("A2:B2", [[username, now.strftime("%Y-%m-%d %H:%M:%S")]])
            return True, ""
        else:
            return False, current_user
    else:
        return False, current_user

def release_lock(username: str):
    df_lock, ws_lock = load_lock_df()
    for i, row in df_lock.iterrows():
        if row["User"] == username:
            ws_lock.update_cell(i + 2, 1, "")
            ws_lock.update_cell(i + 2, 2, "")

# ========== 登出 ==========
def logout():
    st.session_state.clear()
    st.session_state["page"] = "login"
    st.session_state["logged_in"] = False

# ========== 專案編號產生 ==========
def generate_project_number(odm, product_app, cooling):
    odm_code = odm.split(")")[0].strip("(")
    prod_code = product_app.split(")")[0].strip("(")
    cool_code = cooling.split(")")[0].strip("(")
    prefix = f"{odm_code}{prod_code}{cool_code}"

    # 總行數（包含標題列），扣掉 1 才是已經有的資料筆數
    total_rows = len(sheet.get_all_values()) - 1  

    # 下一筆流水號 = 資料筆數 + 1
    new_seq = total_rows + 1  

    return f"{prefix}-{new_seq:03d}"

# ========== 儲存 Google Sheet ==========
def save_to_google_sheet(record):
    record_for_sheet = record.copy()
    record_for_sheet["Project_Number"] = record.get("Project_Number", "")
    record_for_sheet["Spec_Type"] = ", ".join(record.get("Spec_Type", {}).keys())
    # 台灣時間
    record_for_sheet["Update_Time"] = datetime.datetime.now(TAIWAN_TZ).strftime("%Y/%m/%d %H:%M")
    
    row = [record_for_sheet.get(col, "") for col in SHEET_HEADERS]
    sheet.append_row(row)

# ========== 匯出到 Excel 模板 ==========
def export_to_template(record):
    template_path = os.path.join(os.path.dirname(__file__), "Kipo_Project_Form.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    # A. 客戶資訊
    ws["E5"] = record.get("Project_Number", "")
    ws["B5"] = record.get("Sales_User", "")
    ws["B7"] = record.get("ODM_Customers", "")
    ws["E7"] = record.get("Brand_Customers", "")
    ws["B8"] = record.get("Project_Name", "")
    ws["E8"] = record.get("Proposal_Date", "")
    ws["B9"] = record.get("Application_Purpose", "")

    # B. 開案資訊
    ws["B11"] = record.get("Product_Application", "")
    ws["E11"] = record.get("Cooling_Solution", "")
    ws["E13"] = record.get("Delivery_Location", "")
    ws["B12"] = record.get("Sample_Date", "")
    ws["E12"] = record.get("Sample_Qty", "")
    ws["B13"] = record.get("Demand_Qty", "")
    ws["B14"] = record.get("SI", "")
    ws["E14"] = record.get("PV", "")
    ws["B15"] = record.get("MV", "")
    ws["E15"] = record.get("MP", "")

    # ====== C. 規格資訊 ======
    spec_map = {
        "Air Cooling氣冷": "A17",
        "Fan風扇": "C17",
        "Liquid Cooling水冷": "E17"
    }

    # 單位定義表（可自行擴充）
    UNIT_MAP = {
        "Air_Flow": "RPM/Voltage/CFM",
        "Tcase_Max": "°C",
        "Thermal_Resistance": "°C/W",
        "Max_Power": "W",
        "Chip_Length": "mm",
        "Chip_Width": "mm",
        "Chip_Height": "mm",
        "Input_Voltage": "V",
        "Input_Current": "A",
        "Speed": "RPM",
        "Noise": "dB",
        "Sone": "sone",
        "Weight": "g",
        "Cable_Length": "mm",
        "Length": "mm",
        "Width": "mm",
        "Height": "mm",
        "Tj_Max": "°C",
        "T_Inlet": "°C",
        "Flow_Rate": "LPM",
        "Impedance": "KPa",
        "Max_Loading": "lbs"
    }

    specs = record.get("Spec_Type", {})
    if not isinstance(specs, dict):
        specs = {}

    for section, fields in specs.items():
        if section in spec_map:
            lines = [section, ""]

            # Chip/Length/Width/Height 一律排最後
            if section == "Air Cooling氣冷":
                last_keys = ["Chip_Length", "Chip_Width", "Chip_Height"]
            elif section == "Fan風扇":
                last_keys = ["Length", "Width", "Height"]
            elif section == "Liquid Cooling水冷":
                last_keys = ["Chip_Length", "Chip_Width", "Chip_Height"]
            else:
                last_keys = []

            # 先列出一般欄位
            for k, v in fields.items():
                if k not in last_keys:
                    unit = UNIT_MAP.get(k, "")
                    unit_str = f" ({unit})" if unit else ""
                    lines.append(f"{k}{unit_str}: {v}")

            # 再列出最後的尺寸類欄位
            for k in last_keys:
                if k in fields:
                    unit = UNIT_MAP.get(k, "")
                    unit_str = f" ({unit})" if unit else ""
                    lines.append(f"{k}{unit_str}: {fields.get(k, '')}")

            # 寫進對應的格子
            ws[spec_map[section]] = "\n".join(lines)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ========== 頁面：登入 ==========
def login_page():
    st.title("💻 Kipo專案申請系統")
    username = st.text_input("帳號", key="login_username")
    password = st.text_input("密碼", type="password", key="login_password")

    if st.button("🔑 登入"):
        if username in USER_CREDENTIALS and USER_CREDENTIALS[username]["password"] == password:
            st.session_state["logged_in"] = True
            st.session_state["user"] = USER_CREDENTIALS[username]["name"]

            # ✅ 登入後先檢查 Lock
            df_lock, _ = load_lock_df()
            active = df_lock[df_lock["User"] != ""]

            if not active.empty:
                current_user = active.iloc[0]["User"]

                if current_user == st.session_state["user"]:
                    # ✅ 自己持有 Lock → 取屬於自己的「最後一筆」紀錄進入預覽
                    records = sheet.get_all_records()  # list[dict]
                    user = st.session_state["user"]
                    user_records = [r for r in records if str(r.get("Sales_User", "")).strip() == user]

                    if user_records:
                        last = user_records[-1]
                        if isinstance(last.get("Spec_Type"), str):
                            last["Spec_Type"] = {}

                        st.session_state["record"] = last
                        st.session_state["page"] = "preview"
                    else:
                        st.session_state["page"] = "form"
                else:
                    st.session_state["page"] = "form"
            else:
                st.session_state["page"] = "form"
        else:
            st.error("帳號或密碼錯誤，請重新輸入")

# ========== 頁面：A. 客戶資訊 ==========
def render_customer_info():
    st.write(f"### 北辦業務：{st.session_state.get('user','')}")
    st.header("A. 客戶資訊")
    odm = st.selectbox("ODM客戶 (RD)", ["(01)仁寶", "(02)廣達", "(03)緯創", "(04)華勤", "(05)光寶", "(06)技嘉", "(07)智邦", "(08)其他"], key="odm")
    if odm == "(08)其他":
        odm = st.text_input("請輸入ODM客戶", key="odm_other")
    brand = st.selectbox("品牌客戶 (RD)", ["(01)惠普", "(02)聯想", "(03)高通", "(04)華碩", "(05)宏碁", "(06)微星", "(07)技嘉", "(08)其他"], key="brand")
    if brand == "(08)其他":
        brand = st.text_input("請輸入品牌客戶", key="brand_other")
    purpose = st.selectbox("申請目的", ["(01)客戶專案開發", "(02)內部新產品開發", "(03)技術平台預研", "(04)其他"], key="purpose")
    if purpose == "(04)其他":
        purpose = st.text_input("請輸入申請目的", key="purpose_other")
    project_name = st.text_input("客戶專案名稱", key="project_name")
    proposal_date = st.date_input("客戶提案日期", value=datetime.date.today(), key="proposal_date")
    return {"Sales_User": st.session_state["user"], "ODM_Customers": odm, "Brand_Customers": brand,
            "Application_Purpose": purpose, "Project_Name": project_name, "Proposal_Date": proposal_date.strftime("%Y/%m/%d")}


# ========== 頁面：B. 開案資訊 ==========
def render_project_info():
    st.header("B. 開案資訊")
    product_app = st.selectbox("產品應用", ["(01)NB CPU", "(02)NB GPU", "(03)Server", "(04)Automotive(Car)", "(05)Other"], key="product_app")
    if product_app == "(05)Other":
        product_app = st.text_input("請輸入產品應用", key="product_app_other")
    cooling = st.selectbox("散熱方式", ["(01)Air Cooling", "(02)Fan", "(03)Liquid Cooling", "(04)Other"], key="cooling")
    if cooling == "(04)Other":
        cooling = st.text_input("請輸入散熱方式", key="cooling_other")
    delivery = st.selectbox("交貨地點", ["(01)Taiwan", "(02)China", "(03)Thailand", "(04)Vietnam", "(05)Other"], key="delivery")
    if delivery == "(05)Other":
        delivery = st.text_input("請輸入交貨地點", key="delivery_other")
    sample_date = st.date_input("樣品需求日期", value=datetime.date.today(), key="sample_date")
    sample_qty = st.text_input("樣品需求數量", key="sample_qty")
    demand_qty = st.text_input("需求量 (預估數量/總年數)", key="demand_qty")
    col1, col2, col3, col4 = st.columns(4)
    si = col1.text_input("SI", key="si")
    pv = col2.text_input("PV", key="pv")
    mv = col3.text_input("MV", key="mv")
    mp = col4.text_input("MP", key="mp")
    return {"Product_Application": product_app, "Cooling_Solution": cooling, "Delivery_Location": delivery,
            "Sample_Date": sample_date.strftime("%Y/%m/%d"), "Sample_Qty": sample_qty,
            "Demand_Qty": demand_qty, "SI": si, "PV": pv, "MV": mv, "MP": mp}

# ========== 頁面：C. 規格資訊 ==========
def render_spec_info():
    st.header("C. 規格資訊")
    spec_options = st.multiselect("選擇散熱方案", ["Air Cooling氣冷", "Fan風扇", "Liquid Cooling水冷"], key="spec_options")
    spec_data = {}

    if "Air Cooling氣冷" in spec_options:
        st.subheader("Air Cooling氣冷")
        spec_data["Air Cooling氣冷"] = {
            "Air_Flow": st.text_input("Air Flow (RPM/Voltage/CFM)", key="air_flow"),
            "Tcase_Max": st.text_input("Tcase_Max (°C)", key="air_tcase"),
            "Thermal_Resistance": st.text_input("Thermal Resistance (°C/W)", key="air_res"),
            "Max_Power": st.text_input("Max Power (W)", key="air_power"),
            "Chip_Length": st.text_input("Chip_Length (mm)", key="air_len"),
            "Chip_Width": st.text_input("Chip_Width (mm)", key="air_wid"),
            "Chip_Height": st.text_input("Chip_Height (mm)", key="air_hei"),
        }

    if "Fan風扇" in spec_options:
        st.subheader("Fan風扇")
        spec_data["Fan風扇"] = {
            "Max_Power": st.text_input("Max Power (W)", key="fan_power"),
            "Input_Voltage": st.text_input("Input voltage (V)", key="fan_volt"),
            "Input_Current": st.text_input("Input current (A)", key="fan_curr"),
            "PQ": st.text_input("P-Q", key="fan_pq"),
            "Speed": st.text_input("Rotational speed (RPM)", key="fan_speed"),
            "Noise": st.text_input("Noise (dB)", key="fan_noise"),
            "Tone": st.text_input("Tone", key="fan_tone"),
            "Sone": st.text_input("Sone", key="fan_sone"),
            "Weight": st.text_input("Weight (g)", key="fan_weight"),
            "Connector": st.text_input("端子頭型號", key="fan_con"),
            "Wiring": st.text_input("線序", key="fan_wire"),
            "Cable_Length": st.text_input("出框線長", key="fan_cable"),
            "Length": st.text_input("Length (mm)", key="fan_len"),
            "Width": st.text_input("Width (mm)", key="fan_wid"),
            "Height": st.text_input("Height (mm)", key="fan_hei"),
        }

    if "Liquid Cooling水冷" in spec_options:
        st.subheader("Liquid Cooling水冷")
        spec_data["Liquid Cooling水冷"] = {
            "Plate_Form": st.text_input("Plate Form", key="liq_plate"),
            "Max_Power": st.text_input("Max Power (W)", key="liq_max_power"),
            "Tj_Max": st.text_input("Tj_Max (°C)", key="liq_tj"),
            "Tcase_Max": st.text_input("Tcase_Max (°C)", key="liq_tcase"),
            "T_Inlet": st.text_input("T_Inlet (°C)", key="liq_inlet"),
            "Thermal_Resistance": st.text_input("Thermal Resistance (°C/W)", key="liq_res"),
            "Flow_Rate": st.text_input("Flow rate (LPM)", key="liq_flow"),
            "Impedance": st.text_input("Impedance (KPa)", key="liq_imp"),
            "Max_Loading": st.text_input("Max loading (lbs)", key="liq_load"),
            "Chip_Length": st.text_input("Chip_Length (mm)", key="liq_chip_length"),
            "Chip_Width": st.text_input("Chip_Width (mm)", key="liq_chip_width"),
            "Chip_Height": st.text_input("Chip_Height (mm)", key="liq_chip_height"),
        }

    return spec_data

# ========== 頁面：表單 ==========
def form_page():
    if not st.session_state.get("logged_in", False):
        st.session_state["page"] = "login"
        return

    st.title("💻 Kipo專案申請系統")
    if st.button("🚪 登出"): 
        logout()

    customer_info = render_customer_info()
    project_info = render_project_info()
    spec_info = render_spec_info()

    if st.button("✅ 完成"):
        lock_acquired, holder = acquire_lock(st.session_state["user"])
        if not lock_acquired:
            st.warning(f"目前由 {holder} 使用中，請稍後")
            st.warning("當鎖定問題無法透過正常流程解除時，請尋求PM協助處理")
            return

        if any(v in ["", None] for v in customer_info.values()):
            st.error("客戶資訊未完成填寫，請重新確認")
        elif any(v in ["", None] for v in project_info.values()):
            st.error("開案資訊未完成填寫，請重新確認")
        elif not spec_info:
            st.error("規格資訊請至少選擇一種方案")
        else:
            project_number = generate_project_number(
                customer_info["ODM_Customers"], 
                project_info["Product_Application"], 
                project_info["Cooling_Solution"]
            )
            st.session_state["record"] = {
                "Project_Number": project_number, 
                **customer_info, 
                **project_info, 
                "Spec_Type": spec_info
            }
            st.session_state["submitted"] = False
            st.session_state["page"] = "preview"

def preview_page():
    st.title("📋 預覽填寫內容")
    record = st.session_state.get("record", {})
    st.subheader(f"專案編號：{record.get('Project_Number','')}")
    st.write(f"### 北辦業務：{record.get('Sales_User','')}")

    st.subheader("A. 客戶資訊")
    for k, v in {
        "ODM_Customers": "ODM客戶(RD)", 
        "Brand_Customers": "品牌客戶(RD)", 
        "Application_Purpose": "申請目的"
    }.items():
        st.write(f"**{v}：** {record.get(k, '')}")

    st.subheader("B. 開案資訊")
    for k, v in {
        "Product_Application": "產品應用", 
        "Cooling_Solution": "散熱方式", 
        "Delivery_Location": "交貨地點"
    }.items():
        st.write(f"**{v}：** {record.get(k, '')}")

    st.subheader("C. 規格資訊")
    for section, fields in record.get("Spec_Type", {}).items():
        st.markdown(f"**{section}**")
        for k, v in fields.items():
            st.write(f"{k}: {v}")

    col1, col2 = st.columns(2)
    if col1.button("🔙 返回修改"):
        release_lock(st.session_state["user"])
        st.session_state["page"] = "form"

    # 初始化冷卻計時器
    if "last_submit_time" not in st.session_state:
        st.session_state["last_submit_time"] = 0

    now_ts = time.time()
    cooldown = 5  # 冷卻 5 秒

    # ✅ 第一次送出 → 固定專案資料 & 檔名
    if not st.session_state.get("submitted", False):
        if now_ts - st.session_state["last_submit_time"] > cooldown:
            if col2.button("💾 確認送出", key="confirm_submit"):
                st.session_state["last_submit_time"] = now_ts

                save_to_google_sheet(record)
                excel_data = export_to_template(record)
                release_lock(st.session_state["user"])

                apply_date = datetime.datetime.now(TAIWAN_TZ).strftime("%Y%m%d")

                # 🔒 固定資料與檔名
                st.session_state["excel_data"] = excel_data
                st.session_state["fixed_record"] = record.copy()
                st.session_state["fixed_filename"] = f"ProjectForm_{record.get('Project_Number','')}_{apply_date}.xlsx"

                st.session_state["submitted"] = True
                st.success("✅ 申請表單已送出")
        else:
            st.info("⏳ 請稍候再送出，以避免重複紀錄")

    # ✅ 後續下載都用第一次固定的 excel_data & 檔名
    if "excel_data" in st.session_state:
        st.download_button(
            label="⬇️ 自動下載Excel檔案",
            data=st.session_state["excel_data"],
            file_name=st.session_state.get("fixed_filename", "ProjectForm.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ========== 主程式 ==========
def main():
    if "page" not in st.session_state:
        st.session_state["page"] = "login"
        st.session_state["logged_in"] = False

    if st.session_state["page"] == "login":
        login_page()
    elif st.session_state["page"] == "form":
        form_page()
    elif st.session_state["page"] == "preview":
        preview_page()

if __name__ == "__main__":
    main()

#詳細說明請參考以下內容
#------------------------------------------------------------------------------------------------------------------
#1.使用者登入帳號密碼(當帳號密碼輸入正確時，按下登入按鈕後，會轉跳畫面)

#使用者Sam -> 帳號：sam@kipotec.com.tw / 密碼：Kipo-0926969586$$$
#使用者Vivian -> 帳號：sale1@kipotec.com.tw / 密碼：Kipo-0917369466$$$  
#使用者Wendy -> 帳號：sale5@kipotec.com.tw / 密碼：Kipo-0925698417$$$
#使用者Lillian -> 帳號：sale2@kipotec.com.tw / 密碼：Kipo-0905038111$$$

#*登入資訊錯誤會顯示"帳號或密碼輸入錯誤，請重新輸入"，並且會停留在此網頁，直到輸入正確
#------------------------------------------------------------------------------------------------------------------
#2.填寫表單內容(登入成功後，會看到以下選項，當表單內容填寫完後，按下才完成按鈕後，會轉跳畫面)

#<A.客戶資訊>
#北辦業務：顯示使用者名稱(Sam、Vivian、Wendy、Lillian)
#ODM客戶(RD)：下拉式選單(01)仁寶、(02)廣達、(03)緯創、(04)華勤、(05)光寶、(06)技嘉、(07)智邦、(08)其他
#品牌客戶(RD)：下拉式選單(01)惠普、(02)聯想、(03)高通、(04)華碩、(05)宏碁、(06)微星、(07)技嘉、(08)其他
#客戶專案名稱：使用者自行填寫
#客戶提案日期：使用者選擇日期
#申請目的：下拉式選單(01)客戶專案開發、(2)內部新產品開發、(3)技術平台預研 、(4)其他

#<B.開案資訊>
#產品應用：下拉式選單(01)NB CPU、(02)NB GPU、(03)Server、(04)Automotive(Car)、(05)Other
#散熱方式：下拉式選單(01)Air Cooling、(02)Fan、(03)Liquid Cooling、(04)Other
#	樣品需求日期(使用者選擇日期)
#	樣品需求數量 -> 顯示可以打字的地方，使用者自行輸入
#	Schedule
#		SI -> 顯示可以打字的地方，使用者自行輸入
#		PV -> 顯示可以打字的地方，使用者自行輸入
#		MV -> 顯示可以打字的地方，使用者自行輸入
#		MP -> 顯示可以打字的地方，使用者自行輸入
#	需求量(預估數量/總年數) -> 顯示可以打字的地方，使用者自行輸入
#	交貨地點：下拉式選單(01)Taiwan、(02)China、(03)Thailand、(04)Vietnam、(00)Other

#<C.規格資訊>
#Air Cooling：
#	Air Flow(單位 RPM/Voltage/CFM) -> 顯示可以打字的地方，使用者自行輸入
#	Tcase_Max(單位 °C) -> 顯示可以打字的地方，使用者自行輸入
#	Thermal Resistance(單位 ˚C/W) -> 顯示可以打字的地方，使用者自行輸入
#	Max Power(單位 W) -> 顯示可以打字的地方，使用者自行輸入
#	Size L、W、H(單位 mm) -> 顯示可以打字的地方，使用者自行輸入
#Fan：
#	Size L、W、H(單位 mm) -> 顯示可以打字的地方，使用者自行輸入
#	Max Power(單位 W) -> 顯示可以打字的地方，使用者自行輸入
#	Input voltage(單位 V) -> 顯示可以打字的地方，使用者自行輸入
#	Input current(單位 A) -> 顯示可以打字的地方，使用者自行輸入
#	P-Q -> 顯示可以打字的地方，使用者自行輸入
#	Rotational speed(單位 RPM) -> 顯示可以打字的地方，使用者自行輸入
#	Noise(單位 dB) -> 顯示可以打字的地方，使用者自行輸入
#	Tone -> 顯示可以打字的地方，使用者自行輸入
#	Sone -> 顯示可以打字的地方，使用者自行輸入
#	Weight(單位 g) -> 顯示可以打字的地方，使用者自行輸入
#	端子頭型號、線序、出框線長 -> 顯示可以打字的地方，使用者自行輸入
#Liquid Cooling：
#	Plate Form -> 顯示可以打字的地方，使用者自行輸入
#	Max Power(單位 W) -> 顯示可以打字的地方，使用者自行輸入
#	Tj_Max(單位 °C) -> 顯示可以打字的地方，使用者自行輸入
#	Tcase_Max(單位 °C) -> 顯示可以打字的地方，使用者自行輸入
#	T_Inlet(單位 °C) -> 顯示可以打字的地方，使用者自行輸入
#	Chip contact size L、W、H(單位 mm) -> 顯示可以打字的地方，使用者自行輸入
#	Thermal Resistance(單位 °C/W ) -> 顯示可以打字的地方，使用者自行輸入
#	Flow rate(LPM) -> 顯示可以打字的地方，使用者自行輸入
#	Impedance(單位 Kpa) -> 顯示可以打字的地方，使用者自行輸入
#	Max loading(單位 lbs) -> 顯示可以打字的地方，使用者自行輸入

#<D.可行性評估>
#業務主管 -> 顯示可以打字的地方，使用者自行輸入
#研發主管 -> 顯示可以打字的地方，使用者自行輸入

#*C.規格資訊欄位為可複選勾選選項(Air Cooling、Fan、Liquid Cooling)，根據打勾選項會出現對應內容，沒有勾選則不會顯示內容
#*A.客戶資訊和B.開案資訊的內容都要填寫，如果有任一欄位空白時，在按下完成按鈕後，會顯示客戶資訊或開案資訊欄位未填寫完畢，請重新確認的訊息
#------------------------------------------------------------------------------------------------------------------
#3.顯示表單結果(會預覽使用者填寫的表單內容，最下方有下載Excel的按鈕，按下後會自動下載文件，也有取消按鈕，按下後可以回到上一頁修改內容)

#表單最下方會出現以下欄位
#申請日期 -> 顯示可以打字的地方，使用者自行輸入
#專案申請人 -> 顯示可以打字的地方，使用者自行輸入
#填寫日期 -> 顯示可以打字的地方，使用者自行輸入
#規格填寫人 -> 顯示可以打字的地方，使用者自行輸入
#業務主管審核 -> 顯示可以打字的地方，使用者自行輸入
#研發主管審核 -> 顯示可以打字的地方，使用者自行輸入
#核准 -> 顯示可以打字的地方，使用者自行輸入
#------------------------------------------------------------------------------------------------------------------
#4.資料記錄到google sheet

#名稱：Project_Form
#工作頁：Python
#欄位：
#ODM_Customers -> ODM客戶(RD)
#Brand_Customers -> 品牌客戶(RD)
#Application_Purpose -> 申請目的
#Product_Application -> 產品應用
#Cooling_Solution -> 散熱方式
#Delivery_Location -> 交貨地點
#Applicant -> 使用者(Sam、Vivian、Wendy、Lillian)
#Application Deadline -> 系統自行填寫(YYYY/MM/DD/HH/MM)